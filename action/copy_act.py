import os
import sys
sys.path.append(os.getcwd)
import openpyxl as opxl

list_path = os.path.isfile(r"原始檔案路徑")

try: 
    from action.web_get import po_number
    po_path = os.path.isfile('./excel_place/%s.xlsx'%po_number)
except:
    po_path = False

if list_path and po_path:
    from action.web_get import po_time,supplier_number,leadtime,remark
    #po list excel 參數
    list_wb = opxl.load_workbook(r"原始檔案路徑")
    list_ws = list_wb.worksheets[0]
    list_mxr = list_ws.max_row
    list_mxc = list_ws.max_column
    #po excel 參數
    po_wb = opxl.load_workbook('./excel_place/%s.xlsx'%po_number)
    po_ws = po_wb.worksheets[0]
    po_mxr = po_ws.max_row
    po_mxc = po_ws.max_column

    def copy_paste():
        for i in range(2,po_mxr+1,1):
            a01 = list_ws.cell(row=list_mxr+1,column=1,value=po_number)
            a02 = list_ws.cell(row=list_mxr+1,column=2,value=po_time)
            a03 = list_ws.cell(row=list_mxr+1,column=3,value=supplier_number)
            a04 = list_ws.cell(row=list_mxr+1,column=5,value=po_ws.cell(row=i,column=2).value)
            a05 = list_ws.cell(row=list_mxr+1,column=6,value=po_ws.cell(row=i,column=4).value)
            a06 = list_ws.cell(row=list_mxr+1,column=7,value=po_ws.cell(row=i,column=5).value)
            a07 = list_ws.cell(row=list_mxr+1,column=8,value=po_ws.cell(row=i,column=7).value)
            a08 = list_ws.cell(row=list_mxr+1,column=12,value=leadtime)
            a09 = list_ws.cell(row=list_mxr+1,column=15,value=remark)
            list_wb.save(r"原始檔案路徑")
    try:
        list_wb.save(r"原始檔案路徑")
    except:
        print('資料複製貼上失敗\n檔案已被開啟，請關閉 訂單整理.xlsx')
        sys.exit()
    print('資料已複製到 - 訂單整理.xlsx')
else:
    print('找不到訂單整理.xlsx')
    sys.exit()
